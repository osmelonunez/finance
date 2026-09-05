from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
import xlrd

from validators import (
    MAX_RECORD_COMMENT_LENGTH,
    validate_amount,
    validate_concept,
    validate_text_length,
)


MAX_IMPORT_BYTES = 5 * 1024 * 1024
MAX_IMPORT_ROWS = 1000
BANK_FORMATS = {"ing": "ING"}
UNSUPPORTED_CONCEPT_CHARS = re.compile(r"[^A-Za-z0-9ÁÉÍÓÚÜÑáéíóúüñ\s\-\.,:/()&'+]")
HEADERS = ("concept", "amount", "date", "category", "payment_method", "comment")
HEADER_ALIASES = {
    "concept": "concept",
    "concepto": "concept",
    "amount": "amount",
    "importe": "amount",
    "date": "date",
    "fecha": "date",
    "category": "category",
    "categoria": "category",
    "categoría": "category",
    "payment_method": "payment_method",
    "payment method": "payment_method",
    "metodo de pago": "payment_method",
    "método de pago": "payment_method",
    "cuenta/tarjeta": "payment_method",
    "comment": "comment",
    "comentario": "comment",
}


class ImportValidationError(ValueError):
    def __init__(self, errors):
        self.errors = errors
        super().__init__("; ".join(errors))


def build_template(categories=(), payment_methods=()):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Expenses"
    sheet.append(HEADERS)
    sheet.freeze_panes = "A2"
    widths = (28, 14, 14, 22, 26, 36)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width

    lists = workbook.create_sheet("Lists")
    lists.sheet_state = "hidden"
    lists.append(("Categories", "Payment methods"))
    for row_index, value in enumerate(categories, start=2):
        lists.cell(row=row_index, column=1, value=value)
    for row_index, value in enumerate(payment_methods, start=2):
        lists.cell(row=row_index, column=2, value=value)
    if categories:
        validation = DataValidation(
            type="list", formula1=f"Lists!$A$2:$A${len(categories) + 1}", allow_blank=False
        )
        sheet.add_data_validation(validation)
        validation.add(f"D2:D{MAX_IMPORT_ROWS + 1}")
    if payment_methods:
        validation = DataValidation(
            type="list", formula1=f"Lists!$B$2:$B${len(payment_methods) + 1}", allow_blank=True
        )
        sheet.add_data_validation(validation)
        validation.add(f"E2:E{MAX_IMPORT_ROWS + 1}")
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


def _normalise_header(value):
    return HEADER_ALIASES.get(str(value or "").strip().lower())


def _normalise_month(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m")
    text = str(value or "").strip()
    for fmt in ("%Y-%m", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m")
        except ValueError:
            pass
    return None


def _read_rows(data):
    if data.startswith(bytes.fromhex("D0CF11E0")):
        try:
            workbook = xlrd.open_workbook(file_contents=data)
            sheet = workbook.sheet_by_index(0)
            rows = []
            for row_index in range(sheet.nrows):
                row = []
                for col_index in range(sheet.ncols):
                    cell = sheet.cell(row_index, col_index)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        value = xlrd.xldate_as_datetime(value, workbook.datemode)
                    row.append(value)
                rows.append(tuple(row))
            return rows
        except Exception as exc:
            raise ImportValidationError(["The Excel file could not be read."]) from exc
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        workbook.close()
        return rows
    except Exception as exc:
        raise ImportValidationError(["The Excel file could not be read."]) from exc


def preview_bank_statement(upload, bank_format, categories):
    if bank_format not in BANK_FORMATS:
        raise ImportValidationError(["Select a supported bank format."])
    filename = Path(upload.filename or "").name
    if not filename.lower().endswith((".xls", ".xlsx")):
        raise ImportValidationError(["Only .xls and .xlsx Excel files can be imported."])
    data = upload.stream.read(MAX_IMPORT_BYTES + 1)
    if len(data) > MAX_IMPORT_BYTES:
        raise ImportValidationError(["The Excel file exceeds the 5 MB limit."])
    rows = _read_rows(data)
    if bank_format == "ing":
        return _preview_ing(rows, categories)
    raise ImportValidationError(["The selected bank format is not implemented."])


def _preview_ing(rows, categories):
    required = {"f. valor", "categoría", "descripción", "importe (€)"}
    header_index = None
    header_map = {}
    for index, row in enumerate(rows[:20]):
        candidate = {str(value or "").strip().lower(): col for col, value in enumerate(row)}
        if required.issubset(candidate):
            header_index = index
            header_map = candidate
            break
    if header_index is None:
        raise ImportValidationError([
            "The file does not match the ING format. Expected F. VALOR, CATEGORÍA, DESCRIPCIÓN and IMPORTE (€)."
        ])

    category_map = {}
    for category_id, name, display_name in categories:
        category_map[name.casefold()] = category_id
        category_map[display_name.casefold()] = category_id
    parsed = []
    skipped_incomes = 0
    positive_movements = []
    amazon_refunds_by_month = {}
    for source_row, values in enumerate(rows[header_index + 1:], start=header_index + 2):
        if not any(value not in (None, "") for value in values):
            continue
        try:
            amount = Decimal(str(values[header_map["importe (€)"]]).replace(",", "."))
        except (InvalidOperation, TypeError, IndexError):
            raise ImportValidationError([f"Row {source_row}: Amount must be a valid number."])
        if amount >= 0:
            if amount > 0:
                skipped_incomes += 1
                positive_date = _normalise_month(values[header_map["f. valor"]])
                positive_category = str(values[header_map["categoría"]] or "").strip()
                positive_concept = str(values[header_map["descripción"]] or "").strip()
                key = positive_concept.casefold()
                if positive_category.casefold() == "movimientos excluidos":
                    positive_kind = "transfer"
                elif key.startswith("devolución") or key.startswith("devolucion"):
                    positive_kind = "refund"
                else:
                    positive_kind = "income"
                if positive_kind == "refund" and "WWW.AMAZON" in positive_concept.upper():
                    amazon_refunds_by_month.setdefault(positive_date, []).append({
                        "concept": positive_concept,
                        "amount": amount,
                    })
                    skipped_incomes -= 1
                    continue
                clean_concept = " ".join(
                    UNSUPPORTED_CONCEPT_CHARS.sub(" ", positive_concept).split()
                )[:40].rstrip()
                positive_movements.append({
                    "source_row": source_row,
                    "date": positive_date,
                    "category": positive_category,
                    "concept": clean_concept,
                    "amount": amount,
                    "kind": positive_kind,
                    "selected": positive_kind != "transfer",
                })
            continue
        month = _normalise_month(values[header_map["f. valor"]])
        if not month:
            raise ImportValidationError([f"Row {source_row}: Date is not valid."])
        bank_category = str(values[header_map["categoría"]] or "").strip()
        source_description = str(values[header_map["descripción"]] or "").strip()
        original_description = source_description
        if original_description.casefold().startswith("pago en "):
            original_description = original_description[len("Pago en "):].lstrip()
        description_key = original_description.upper()
        target_category = None
        if "WWW.AMAZON" in description_key:
            original_description = "Amazon"
            target_category = "Home"
        elif "AHORRAMAS" in description_key:
            original_description = "AHORRAMAS"
            target_category = "Food"
        elif "UBR" in description_key and "PENDING.UBER.COM" in description_key:
            original_description = "Uber"
            target_category = "Transport"
        elif "METRO DE MADRID" in description_key:
            target_category = "Transport"
        elif "BILL SENCILLO AVANZA CRTMMADRID" in description_key:
            target_category = "Transport"
        elif "FARMACIA" in description_key:
            target_category = "Health"
        elif any(
            merchant in description_key
            for merchant in ("PIZZERIA NAPOLITANA", "SANTAGLORIA", "STARBUCKS", "CAFETERIA VIPS")
        ):
            target_category = "Leisure"
        elif "APPLE.COMBILL" in description_key:
            original_description = "iCloud"
            target_category = "Subscriptions"
        elif "EMPRESA MUNICIPAL EMT MADRID" in description_key:
            target_category = "Transport"
        elif "PAYPAL" in description_key:
            original_description = "PayPal"
        description = " ".join(UNSUPPORTED_CONCEPT_CHARS.sub(" ", original_description).split())
        category_id = (
            category_map.get(target_category.casefold())
            if target_category
            else category_map.get(bank_category.casefold())
        )
        warnings = []
        if category_id is None:
            warnings.append("Category needs mapping")
        if len(description) > 40:
            warnings.append("Description shortened to 40 characters")
            description = description[:40].rstrip()
        _, error = validate_concept(description)
        if error:
            warnings.append(error)
        parsed.append({
            "source_row": source_row,
            "date": month,
            "concept": description,
            "amount": abs(amount),
            "bank_category": bank_category,
            "category_id": category_id,
            "warnings": warnings,
            "selected": bank_category.casefold() != "movimientos excluidos",
            "source_description": source_description,
        })
        if len(parsed) > MAX_IMPORT_ROWS:
            raise ImportValidationError([f"The limit is {MAX_IMPORT_ROWS} expenses per import."])
    if not parsed and not positive_movements:
        raise ImportValidationError(["The Excel file does not contain movements to import."])
    grouped = []
    amazon_by_month = {}
    for row in parsed:
        if row["concept"].casefold() != "amazon":
            grouped.append(row)
            continue
        existing = amazon_by_month.get(row["date"])
        if existing is None:
            row["grouped_count"] = 1
            row["grouped_items"] = [{
                "concept": row["source_description"],
                "amount": row["amount"],
            }]
            amazon_by_month[row["date"]] = row
            grouped.append(row)
            continue
        existing["amount"] += row["amount"]
        existing["grouped_count"] += 1
        existing["grouped_items"].append({
            "concept": row["source_description"],
            "amount": row["amount"],
        })
        existing["selected"] = existing["selected"] or row["selected"]
        existing["warnings"] = list(dict.fromkeys(existing["warnings"] + row["warnings"]))
    for month, refunds in amazon_refunds_by_month.items():
        amazon_group = amazon_by_month.get(month)
        if amazon_group is None:
            continue
        for refund in refunds:
            amazon_group["amount"] -= refund["amount"]
            amazon_group["grouped_count"] += 1
            amazon_group["grouped_items"].append({
                "concept": refund["concept"],
                "amount": -refund["amount"],
            })
    grouped = [row for row in grouped if row["amount"] > 0]
    return grouped, skipped_incomes, positive_movements


def parse_workbook(upload, categories, payment_methods, movement_type="expense"):
    if movement_type not in {"expense", "income"}:
        raise ValueError("Unsupported movement type.")
    filename = Path(upload.filename or "").name
    if not filename.lower().endswith(".xlsx"):
        raise ImportValidationError(["Only .xlsx Excel files can be imported."])
    data = upload.stream.read(MAX_IMPORT_BYTES + 1)
    if len(data) > MAX_IMPORT_BYTES:
        raise ImportValidationError(["The Excel file exceeds the 5 MB limit."])
    try:
        workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        raw_headers = next(iterator, None)
    except Exception as exc:
        raise ImportValidationError(["The Excel file could not be read."]) from exc

    headers = [_normalise_header(value) for value in (raw_headers or ())]
    required = {"concept", "amount", "date"}
    if movement_type == "expense":
        required.add("category")
    missing = sorted(required - set(headers))
    if missing:
        raise ImportValidationError([f"Missing required columns: {', '.join(missing)}."])
    if len([header for header in headers if header]) != len(set(header for header in headers if header)):
        raise ImportValidationError(["The Excel file contains duplicate columns."])

    category_map = {name.casefold(): (category_id, name) for category_id, name in categories}
    payment_map = {name.casefold(): (method_id, name) for method_id, name in payment_methods}
    parsed = []
    errors = []
    for row_number, values in enumerate(iterator, start=2):
        row = {header: values[index] if index < len(values) else None for index, header in enumerate(headers) if header}
        if not any(value not in (None, "") for value in row.values()):
            continue
        if len(parsed) >= MAX_IMPORT_ROWS:
            errors.append(f"Row {row_number}: the limit is {MAX_IMPORT_ROWS} records per import.")
            break

        row_errors = []
        concept, error = validate_concept(str(row.get("concept") or ""))
        if error:
            row_errors.append(error)
        amount, error = validate_amount(str(row.get("amount") or ""))
        if error:
            row_errors.append(error)
        month = _normalise_month(row.get("date"))
        if not month:
            row_errors.append("Date must use YYYY-MM or be a valid Excel date.")

        category_id = None
        category_name = str(row.get("category") or "").strip()
        if movement_type == "expense":
            category = category_map.get(category_name.casefold())
            if not category:
                row_errors.append(f"Category '{category_name}' does not exist.")
            else:
                category_id = category[0]

        payment_method_id = None
        payment_name = str(row.get("payment_method") or "").strip()
        if movement_type == "expense" and payment_name:
            payment = payment_map.get(payment_name.casefold())
            if not payment:
                row_errors.append(f"Active account/card '{payment_name}' does not exist.")
            else:
                payment_method_id = payment[0]

        comment, error = validate_text_length(
            str(row.get("comment") or ""), "Comment", MAX_RECORD_COMMENT_LENGTH
        )
        if error:
            row_errors.append(error)
        if row_errors:
            errors.extend(f"Row {row_number}: {message}" for message in row_errors)
            continue
        parsed.append({
            "concept": concept,
            "amount": amount,
            "date": month,
            "category_id": category_id,
            "payment_method_id": payment_method_id,
            "comment": comment,
        })
    workbook.close()
    if not parsed and not errors:
        errors.append("The Excel file does not contain records to import.")
    if errors:
        raise ImportValidationError(errors)
    return parsed
